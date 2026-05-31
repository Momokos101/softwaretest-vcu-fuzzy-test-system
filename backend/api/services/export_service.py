"""V2 export service: JSON, CSV, Excel, and bq_new payloads."""
from __future__ import annotations

import io
import json
import csv
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from api.models.schemas import ExportFormat, ExportRequest
from api.services import coverage_service, performance_service, requirement_service, risk_service, test_design_service


def export_data(request: ExportRequest) -> tuple[bytes, str, str]:
    data = collect_data(request)
    if request.format == ExportFormat.JSON:
        return _export_json(data)
    if request.format == ExportFormat.CSV:
        return _export_csv(data)
    return _export_excel(data)


def collect_data(request: ExportRequest) -> dict[str, Any]:
    selected_ids = set(request.requirement_ids or [])

    def selected(req_id: str) -> bool:
        return not selected_ids or req_id in selected_ids

    requirements = [item for item in requirement_service.get_all_requirements() if selected(item.id)]
    parsed = [item for item in requirement_service.get_all_parsed_requirements() if selected(item.requirement_id)]
    risks = [item for item in risk_service.get_all_risk_results() if selected(item.requirement_id)]
    coverage = [item for item in coverage_service.list_coverage_items() if selected(item.requirement_id)]
    strategies = [item for item in coverage_service.list_strategies() if selected(item.requirement_id)]
    cases = [item for item in test_design_service.get_all_test_cases() if selected(item.requirement_id)]
    cases = _filter_cases_by_scope(cases, request)

    data: dict[str, Any] = {}
    if request.scope.include_requirements:
        data["requirements"] = [item.model_dump(mode="json") for item in requirements]
    if request.scope.include_parsed_requirements:
        data["parsed_requirements"] = [item.model_dump(mode="json") for item in parsed]
    if request.scope.include_risk_analysis:
        data["risk_analysis"] = [item.model_dump(mode="json") for item in risks]
    if request.scope.include_coverage_items:
        data["coverage_items"] = [item.model_dump(mode="json") for item in coverage]
    if request.scope.include_strategies:
        data["strategies"] = [item.model_dump(mode="json") for item in strategies]
    if request.scope.include_test_cases:
        rows = [item.model_dump(mode="json") for item in cases]
        if not request.scope.include_execution_results:
            for row in rows:
                row["execution_result"] = None
        data["test_cases"] = rows
    if request.scope.include_bq_new_cases:
        data["bq_new_cases"] = [_to_bq_new(item) for item in cases]
    if request.scope.include_traceability_matrix:
        data["traceability_matrix"] = _traceability(requirements, risks, coverage, cases)
    data["performance"] = performance_service.summary()
    return data


def _filter_cases_by_scope(cases, request: ExportRequest):
    allowed = set()
    if request.scope.include_ep_cases:
        allowed.add("EP")
    if request.scope.include_bva_cases:
        allowed.add("BVA")
    if request.scope.include_dt_cases:
        allowed.add("DT")
    if request.scope.include_st_cases:
        allowed.add("ST")
    if request.scope.include_scenario_cases:
        allowed.add("SC")
    return [case for case in cases if not allowed or case.technique.value in allowed]


def _traceability(requirements, risks, coverage, cases) -> list[dict[str, Any]]:
    risk_by_req = {item.requirement_id: item for item in risks}
    coverage_by_req: dict[str, list] = {}
    cases_by_req: dict[str, list] = {}
    for item in coverage:
        coverage_by_req.setdefault(item.requirement_id, []).append(item)
    for item in cases:
        cases_by_req.setdefault(item.requirement_id, []).append(item)

    rows = []
    for req in requirements:
        risk = risk_by_req.get(req.id)
        rows.append(
            {
                "requirement_id": req.id,
                "title": req.title,
                "module": req.module,
                "risk_rpn": risk.rpn if risk else None,
                "extent": risk.extent if risk else None,
                "coverage_item_count": len(coverage_by_req.get(req.id, [])),
                "test_case_count": len(cases_by_req.get(req.id, [])),
                "test_case_ids": ",".join(case.id for case in cases_by_req.get(req.id, [])),
            }
        )
    return rows


def _to_bq_new(case) -> dict[str, Any]:
    return {
        "type": case.type,
        "in_data": [item.model_dump(mode="json", exclude_none=True) for item in case.in_data],
        "expected_results": [item.model_dump(mode="json") for item in case.expected_results],
        "error": [item.model_dump(mode="json") for item in case.error],
        "est_time": case.est_time,
    }


def _export_json(data: dict[str, Any]) -> tuple[bytes, str, str]:
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"), "application/json", "autotestdesign_v2_export.json"


def _export_csv(data: dict[str, Any]) -> tuple[bytes, str, str]:
    output = io.StringIO()
    target = "traceability_matrix" if data.get("traceability_matrix") else "test_cases"
    rows = data.get(target) or []
    if rows:
        headers = list(rows[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: _stringify(row.get(header)) for header in headers})
    return output.getvalue().encode("utf-8-sig"), "text/csv", "autotestdesign_v2_export.csv"


def _export_excel(data: dict[str, Any]) -> tuple[bytes, str, str]:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in data.items():
        if isinstance(rows, list) and rows:
            _add_sheet(wb, sheet_name[:31], rows)
    if not wb.sheetnames:
        _add_sheet(wb, "Export", [{"message": "No data selected"}])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "autotestdesign_v2_export.xlsx"


def _add_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(name)
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([_stringify(row.get(header)) for header in headers])
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 80)


def _stringify(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value
